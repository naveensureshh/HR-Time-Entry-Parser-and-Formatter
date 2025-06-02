
import os
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from graph_utils import download_file, upload_file  # Uses Graph API helpers

def parse_scheduled_start(scheduled: str):
    if not scheduled or scheduled.lower() in ('off', ''):
        return None
    start_str = scheduled.split('-')[0].strip()
    fmt = "%I%p" if ':' not in start_str else "%I:%M%p"
    try:
        return datetime.strptime(start_str, fmt).time()
    except ValueError:
        return None

def load_timesheet(stream: BytesIO, dt_format: str) -> pd.DataFrame:
    df = pd.read_csv(stream)
    df.columns = df.columns.str.strip()
    drop_cols = [c for c in df.columns if c.lower().startswith(('legend', 'unnamed'))]
    df = df.drop(columns=drop_cols, errors='ignore')
    df['Employee Name'] = df['Employee Name'].astype(str).str.strip()
    df['NameClean'] = (
        df['Employee Name']
        .str.replace(r'\s*\(.*?\)', '', regex=True)
        .str.replace(r'\s+', ' ', regex=True)
        .str.strip()
        .str.lower()
    )
    df['Clock-in Time'] = pd.to_datetime(df['Clock-in Time'], format=dt_format, errors='coerce')
    df['Date'] = df['Clock-in Time'].dt.date
    print("🔍 Timesheet data preview:")
    print(df.head())
    return df

def normalize_schedule(schedule_df: pd.DataFrame) -> pd.DataFrame:
    sched = schedule_df.copy()
    sched.columns = sched.columns.str.strip()
    sched['NameClean'] = (
        sched['Name'].astype(str)
        .str.replace(r'\s*\(.*?\)', '', regex=True)
        .str.replace(r'\s+', ' ', regex=True)
        .str.strip()
        .str.lower()
    )
    return sched

def analyze_attendance(schedule_df: pd.DataFrame,
                       timesheet_df: pd.DataFrame,
                       target_date) -> pd.DataFrame:
    target = pd.to_datetime(target_date).date()
    sched = normalize_schedule(schedule_df)
    ts = timesheet_df.copy()
    ts_today = ts[ts['Date'] == target]
    grouped = ts_today.groupby('NameClean')
    weekday = target.strftime('%A')

    rows = []
    seen = set()

    for _, r in sched.iterrows():
        orig = r['Name']
        norm = r['NameClean']
        sched_str = str(r.get(weekday, '')).strip()
        if not sched_str or sched_str.lower() == 'off':
            status, clk = 'Not Scheduled', ''
        else:
            start_time = parse_scheduled_start(sched_str)
            if norm in grouped.groups:
                first_in = grouped.get_group(norm)['Clock-in Time'].min().time()
                late = start_time and first_in > start_time
                status = 'Late Clock-in' if late else 'Present'
                clk = first_in.strftime('%I:%M %p')
            else:
                status, clk = 'Absent', ''
        rows.append((orig, norm, status, sched_str, clk, target))
        seen.add(norm)

    extras = set(ts_today['NameClean']) - set(sched['NameClean'])
    for extra in sorted(extras):
        if extra in seen:
            continue
        group = grouped.get_group(extra)
        orig = group['Employee Name'].iloc[0]
        time_in = group['Clock-in Time'].min().time()
        rows.append((orig, extra, 'Present (No Schedule)', '', time_in.strftime('%I:%M %p'), target))

    df_out = pd.DataFrame(rows, columns=[
        'Employee Name', 'NameClean', 'Status', 'Scheduled', 'Clock-in', 'Date'
    ])
    df_out = df_out.drop_duplicates(subset=['NameClean'], keep='first')
    df_out = df_out.drop(columns='NameClean')
    df_out['Employee Name'] = df_out['Employee Name'].str.title()
    print("📊 Attendance summary preview:")
    print(df_out.head())
    return df_out.sort_values(['Status', 'Employee Name']).reset_index(drop=True)

def create_summary_excel_in_memory(df: pd.DataFrame) -> BytesIO:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        ws = writer.book.active
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
        ws.freeze_panes = 'A2'
    buffer.seek(0)
    return buffer

def main():
    print("🚀 Starting HR attendance automation...")

    # Download inputs from SharePoint
    print("⬇️ Downloading reference schedule...")
    schedule_stream = download_file(match_name="cleaned_dataset.csv")

    print("⬇️ Downloading latest timesheet...")
    timesheet_stream = download_file(latest_csv_only=True)

    # Load and process data
    schedule_df = pd.read_csv(schedule_stream)
    timesheet_df = load_timesheet(timesheet_stream, "%m/%d/%Y %I:%M %p")

    dates = timesheet_df['Date'].dropna().unique()
    if len(dates) != 1:
        raise ValueError(f"Expected 1 date in timesheet, found: {dates}")
    target = dates[0]

    summary_df = analyze_attendance(schedule_df, timesheet_df, target)
    excel_stream = create_summary_excel_in_memory(summary_df)

    # Upload summary back to SharePoint
    out_filename = f"attendance_summary_{target}.xlsx"
    print(f"⬆️ Uploading summary: {out_filename}")
    result = upload_file(excel_stream, out_filename)
    print(f"✅ Uploaded to SharePoint: {result.get('webUrl')}")

if __name__ == "__main__":
    main()
